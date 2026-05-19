#!/usr/bin/env python3
"""Build word + default prompt rows for words that are in word-sets.json but
not in _library/, and not already on image-prompts.xlsx (non-highlighted).

Includes minimal-pair sets (``pairs``) and single-word sets (``words``).

Writes (same two columns as image-prompts.xlsx: **word**, **prompt** only —
nothing else, so nothing can be mistaken for a prompt):

  tools/missing-not-in-prompts-with-prompts.xlsx
  tools/missing-not-in-prompts-with-prompts.csv
  tools/missing-not-in-prompts-xlsx.csv   (legacy filename; same two columns)

The image batch reads only the **prompt** column for OpenAI. To generate:

    python3 tools/batch_generate_openai.py --csv tools/missing-not-in-prompts-with-prompts.xlsx

Re-run this script after you change word-sets or the library:

    python3 tools/generate_missing_prompt_sheet.py
"""

from __future__ import annotations

import csv
import json
import sys
from pathlib import Path
from typing import Dict, List, Set, Tuple


def find_repo_root(start: Path) -> Path:
    here = start.resolve()
    for candidate in [here, *here.parents]:
        if (candidate / "word-lists").is_dir() and (candidate / "word-images").is_dir():
            return candidate
    return start.resolve()


REPO_ROOT = find_repo_root(Path(__file__).parent)


def cell_is_highlighted(cell) -> bool:
    try:
        fill = cell.fill
        if not fill or not getattr(fill, "patternType", None) or fill.patternType == "none":
            return False
        color = fill.fgColor
        ctype = getattr(color, "type", None)
        if ctype == "theme":
            theme = getattr(color, "theme", None)
            tint = getattr(color, "tint", None) or 0.0
            if (theme == 0 or theme is None) and abs(tint) < 1e-6:
                return False
            return True
        if ctype == "rgb":
            rgb = getattr(color, "rgb", None)
            if isinstance(rgb, str) and rgb.upper() not in {"00000000", "FFFFFFFF"}:
                return True
            return False
        if ctype == "indexed":
            return getattr(color, "indexed", None) not in (None, 64, 65)
        return True
    except Exception:
        return False


def load_xlsx_words(path: Path) -> Set[str]:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    header_cells = next(ws.iter_rows(max_row=1), None)
    if not header_cells:
        return set()
    header = [(str(c.value).strip().lower() if c.value else "") for c in header_cells]
    idx_word = header.index("word")
    out: Set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=False):
        if not row or len(row) <= idx_word:
            continue
        if any(cell_is_highlighted(c) for c in row):
            continue
        v = row[idx_word].value
        if v:
            out.add(str(v).strip().lower())
    return out


SUFFIX = (
    "centered, isolated, plain white background, minimal style, bold outline, "
    "no text, square composition"
)


def default_prompt(word: str) -> str:
    """Starter prompt aligned with your usual image-prompts.xlsx style."""
    w = word.strip()
    if not w:
        return f"clean vector clipart of {word}, {SUFFIX}"
    # Single letter (minimal-pair letter cards)
    if len(w) == 1 and w.isalpha():
        return (
            f"clean vector clipart of the letter {w.upper()}, {SUFFIX}"
        )
    return f"clean vector clipart of {w}, {SUFFIX}"


def iter_words_in_set(word_set: dict):
    """Yield each word from a set (minimal pairs or single-word lists)."""
    set_type = (word_set.get("setType") or "pairs").strip().lower()
    if set_type == "single":
        for raw in word_set.get("words", []) or []:
            w = str(raw).strip()
            if w:
                yield w
        return
    for pair in word_set.get("pairs", []) or []:
        for raw in pair:
            w = str(raw).strip()
            if w:
                yield w


def collect_words_from_sets(sets: list) -> Dict[str, str]:
    lower_to_display: Dict[str, str] = {}
    for s in sets:
        for word in iter_words_in_set(s):
            wl = word.lower()
            if wl not in lower_to_display:
                lower_to_display[wl] = word
    return lower_to_display


def main() -> int:
    word_sets_path = REPO_ROOT / "word-lists" / "word-sets.json"
    library_dir = REPO_ROOT / "word-images" / "_library"
    prompt_xlsx = REPO_ROOT / "tools" / "image-prompts.xlsx"
    out_xlsx = REPO_ROOT / "tools" / "missing-not-in-prompts-with-prompts.xlsx"
    out_csv = REPO_ROOT / "tools" / "missing-not-in-prompts-with-prompts.csv"
    legacy_name_csv = REPO_ROOT / "tools" / "missing-not-in-prompts-xlsx.csv"

    lib_stems: Set[str] = set()
    if library_dir.is_dir():
        for p in library_dir.glob("*.png"):
            k = p.stem.strip().lower()
            if k:
                lib_stems.add(k)

    with word_sets_path.open(encoding="utf-8") as f:
        sets = json.load(f)
    pair_sets = sum(
        1 for s in sets if (s.get("setType") or "pairs").strip().lower() != "single"
    )
    single_sets = sum(
        1 for s in sets if (s.get("setType") or "").strip().lower() == "single"
    )

    lower_to_display = collect_words_from_sets(sets)
    missing_lower = sorted(set(lower_to_display) - lib_stems, key=lambda k: lower_to_display[k].lower())

    in_xlsx: Set[str] = set()
    if prompt_xlsx.is_file():
        in_xlsx = load_xlsx_words(prompt_xlsx)

    rows: List[Tuple[str, str]] = []
    for wl in missing_lower:
        if wl in in_xlsx:
            continue
        disp = lower_to_display[wl]
        rows.append((disp, default_prompt(disp)))

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "image-prompts"
    ws.append(["word", "prompt"])
    for word, prompt in rows:
        ws.append([word, prompt])
    wb.save(out_xlsx)

    with out_csv.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["word", "prompt"])
        w.writerows(rows)

    with legacy_name_csv.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["word", "prompt"])
        w.writerows(rows)

    print(f"Word sets: {pair_sets} pair, {single_sets} single-word.")
    print(f"Unique words in word-sets.json: {len(lower_to_display)}")
    print(f"Rows written: {len(rows)}")
    print(f"XLSX: {out_xlsx}")
    print(f"CSV:  {out_csv}")
    print(f"CSV:  {legacy_name_csv}  (same content; name you already use)")
    print("\nDefault prompt pattern:")
    print(f'  clean vector clipart of <word>, {SUFFIX}')
    print("  (single letters use: clean vector clipart of the letter X, ...)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
