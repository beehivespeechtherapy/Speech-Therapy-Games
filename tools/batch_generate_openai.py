#!/usr/bin/env python3
"""Bulk-generate clipart images via OpenAI's ``gpt-image-1`` model.

Reads a CSV (or .xlsx) of word prompts and writes PNGs into the right
per-folder spots under ``word-images/`` AND into the shared
``word-images/_library/`` folder.

Typical usage from the project root:

    # Dry run (no API calls) to preview what would be generated
    python3 tools/batch_generate_openai.py --dry-run

    # Real run — slow and safe (1 image at a time, 12s between calls)
    python3 tools/batch_generate_openai.py

    # Faster with retries (use only if your account allows it)
    python3 tools/batch_generate_openai.py --concurrency 2 --min-seconds-between-jobs 6

Requires:
    pip3 install openai python-dotenv openpyxl

Set OPENAI_API_KEY in a ``.env`` file in the project root, e.g.:
    OPENAI_API_KEY=sk-...
"""
from __future__ import annotations

import argparse
import base64
import csv
import json
import os
import random
import re
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Set, Tuple


# ---- Setup helpers ----------------------------------------------------------


def find_repo_root(start: Path) -> Path:
    here = start.resolve()
    for candidate in [here, *here.parents]:
        if (candidate / "word-lists").is_dir() and (candidate / "word-images").is_dir():
            return candidate
    return start.resolve()


REPO_ROOT = find_repo_root(Path(__file__).parent)


def load_dotenv_simple(path: Path) -> None:
    """A tiny .env loader so we don't strictly need python-dotenv installed."""
    if not path.is_file():
        return
    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        if "=" not in line:
            continue
        key, _, value = line.partition("=")
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


# ---- Word-sets / folder lookup ---------------------------------------------


def load_word_sets(path: Path) -> List[dict]:
    if not path.is_file():
        raise SystemExit(f"word-sets.json not found at: {path}")
    with path.open(encoding="utf-8") as handle:
        data = json.load(handle)
    if not isinstance(data, list):
        raise SystemExit("word-sets.json must be a JSON array.")
    return data


def word_to_folders(sets: List[dict]) -> Dict[str, List[str]]:
    """Return {lowercase_word: [folder_name, ...]} — every set folder that word
    appears in, so we can mirror the generated image into every needed folder.
    """
    mapping: Dict[str, Set[str]] = {}
    for s in sets:
        folder = (s.get("folder") or "").strip()
        if not folder:
            continue
        for pair in s.get("pairs", []) or []:
            for raw in pair:
                if not raw:
                    continue
                word = str(raw).strip()
                if not word:
                    continue
                mapping.setdefault(word.lower(), set()).add(folder)
    return {w: sorted(folders) for w, folders in mapping.items()}


# ---- CSV / XLSX prompt loading ---------------------------------------------


def load_prompt_rows_from_csv(path: Path) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    with path.open(encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            word = (row.get("word") or "").strip()
            prompt = (row.get("prompt") or "").strip()
            folder = (row.get("folder") or "").strip()
            if not word:
                continue
            rows.append({"folder": folder, "word": word, "prompt": prompt})
    return rows


def cell_is_highlighted(cell) -> bool:
    """Return True if an openpyxl cell has a non-default fill color set.

    Detects:
      * RGB fills (any non-white, non-empty RGB)
      * Theme fills (e.g., yellow highlight = theme 5, tint 0.6 on macOS Excel)
      * Indexed-color fills
    Treats theme 0 / tint 0 / rgb 00000000 / FFFFFFFF as 'no highlight'.
    """
    try:
        fill = cell.fill
        if not fill:
            return False
        pattern = getattr(fill, "patternType", None)
        if not pattern or pattern == "none":
            return False
        color = getattr(fill, "fgColor", None)
        if color is None:
            return False
        ctype = getattr(color, "type", None)
        if ctype == "rgb":
            rgb = getattr(color, "rgb", None)
            if isinstance(rgb, str) and rgb.upper() not in {"00000000", "FFFFFFFF"}:
                return True
            return False
        if ctype == "theme":
            # Theme 0 + tint 0 is the default "no fill" representation in some files.
            theme = getattr(color, "theme", None)
            tint = getattr(color, "tint", None) or 0.0
            if (theme == 0 or theme is None) and abs(tint) < 1e-6:
                return False
            return True
        if ctype == "indexed":
            indexed = getattr(color, "indexed", None)
            return indexed not in (None, 64, 65)  # 64/65 = system default
        return True
    except Exception:
        return False


def load_prompt_rows_from_xlsx(path: Path, skip_highlighted: bool = True
                               ) -> List[Dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit(
            "Reading .xlsx requires openpyxl. Install it with:\n"
            "    pip3 install openpyxl"
        )
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=False)
    header_cells = next(rows_iter, None)
    if not header_cells:
        return []
    header = [(str(c.value).strip().lower() if c.value else "") for c in header_cells]
    try:
        idx_word = header.index("word")
        idx_prompt = header.index("prompt")
    except ValueError:
        raise SystemExit(
            f"{path} must have at least 'word' and 'prompt' columns in row 1."
        )
    if "folder" in header:
        idx_folder = header.index("folder")
    elif "example_folder" in header:
        idx_folder = header.index("example_folder")
    else:
        idx_folder = None

    rows: List[Dict[str, str]] = []
    skipped_highlighted = 0
    for row in rows_iter:
        if not row:
            continue
        if skip_highlighted and any(cell_is_highlighted(c) for c in row):
            skipped_highlighted += 1
            continue
        word_cell = row[idx_word] if len(row) > idx_word else None
        if not word_cell or word_cell.value is None:
            continue
        word = str(word_cell.value).strip()
        if not word:
            continue
        prompt_cell = row[idx_prompt] if len(row) > idx_prompt else None
        prompt = (str(prompt_cell.value).strip()
                  if prompt_cell is not None and prompt_cell.value else "")
        folder = ""
        if idx_folder is not None and len(row) > idx_folder:
            folder_cell = row[idx_folder]
            if folder_cell.value:
                folder = str(folder_cell.value).strip()
        rows.append({"folder": folder, "word": word, "prompt": prompt})
    if skip_highlighted and skipped_highlighted:
        print(f"Skipped {skipped_highlighted} highlighted row(s) in {path.name}.")
    return rows


def load_prompt_rows(path: Path, skip_highlighted: bool = True) -> List[Dict[str, str]]:
    if path.suffix.lower() == ".xlsx":
        return load_prompt_rows_from_xlsx(path, skip_highlighted=skip_highlighted)
    return load_prompt_rows_from_csv(path)


# ---- Prompt building -------------------------------------------------------


DEFAULT_TEMPLATE_PREFIX = (
    "Friendly clean vector clipart of"
)
DEFAULT_TEMPLATE_SUFFIX = (
    ", centered, isolated on a plain white background, minimal style, "
    "bold outline, no text, square composition. Suitable for a "
    "children's speech therapy app."
)

_BOILERPLATE_PATTERNS = [
    r"clean vector clipart of",
    r"clipart of",
    r"centered,?",
    r"isolated,?",
    r"plain white background",
    r"minimal style,?",
    r"bold outline,?",
    r"no text,?",
    r"square composition\.?",
]


def article_for(word: str) -> str:
    if not word:
        return "a"
    return "an" if word[0].lower() in "aeiou" else "a"


def extract_subject_from_prompt(prompt: str, word: str) -> Optional[str]:
    """If the prompt contains a customized phrase beyond our boilerplate,
    return it. Otherwise return None.
    """
    if not prompt:
        return None
    s = prompt
    for pat in _BOILERPLATE_PATTERNS:
        s = re.sub(pat, " ", s, flags=re.IGNORECASE)
    # remove the bare word itself
    s = re.sub(rf"\b{re.escape(word)}\b", " ", s, flags=re.IGNORECASE)
    s = re.sub(r"\s+", " ", s).strip(" ,.;:-")
    if not s:
        return None
    if len(s) < 3:
        return None
    return s


def build_final_prompt(word: str, raw_prompt: str, mode: str) -> str:
    """Build the prompt actually sent to OpenAI."""
    word_clean = word.strip()
    art = article_for(word_clean)

    if mode == "csv":
        # Use the CSV's prompt verbatim
        return raw_prompt or f"{DEFAULT_TEMPLATE_PREFIX} {art} {word_clean}{DEFAULT_TEMPLATE_SUFFIX}"

    if mode == "template":
        return f"{DEFAULT_TEMPLATE_PREFIX} {art} {word_clean}{DEFAULT_TEMPLATE_SUFFIX}"

    # mode == "merge": pull custom subject phrase if present, else default
    subject = extract_subject_from_prompt(raw_prompt, word_clean)
    if subject:
        return (
            f"{DEFAULT_TEMPLATE_PREFIX} {subject} (illustrating the word "
            f"'{word_clean}'){DEFAULT_TEMPLATE_SUFFIX}"
        )
    return f"{DEFAULT_TEMPLATE_PREFIX} {art} {word_clean}{DEFAULT_TEMPLATE_SUFFIX}"


# ---- Image saving ----------------------------------------------------------


def destination_paths(word: str, folder_hint: str, folder_map: Dict[str, List[str]],
                      word_images_root: Path, library_dir: Path,
                      mirror_sets: bool) -> List[Path]:
    """Where the PNG bytes should be written. Always includes the library, plus
    every set folder the word appears in (when mirror_sets is True)."""
    paths: List[Path] = []
    library_dir.mkdir(parents=True, exist_ok=True)
    paths.append(library_dir / f"{word}.png")
    if mirror_sets:
        folders = folder_map.get(word.lower(), [])
        if folder_hint and folder_hint not in folders:
            folders = [folder_hint, *folders]
        for folder in folders:
            dest = word_images_root / folder
            dest.mkdir(parents=True, exist_ok=True)
            paths.append(dest / f"{word}.png")
    return paths


def write_png_bytes(paths: Iterable[Path], data: bytes) -> int:
    count = 0
    for p in paths:
        p.parent.mkdir(parents=True, exist_ok=True)
        with p.open("wb") as fh:
            fh.write(data)
        count += 1
    return count


# ---- API call --------------------------------------------------------------


def is_transient_network_error(exc: Exception) -> bool:
    text = (repr(exc) + " " + str(exc)).lower()
    needles = [
        "connection reset",
        "read operation timed out",
        "timed out",
        "remoteprotocolerror",
        "remotedisconnected",
        "temporary failure in name resolution",
        "nodename nor servname",
        "broken pipe",
        "ssl",
        "incomplete read",
    ]
    return any(n in text for n in needles)


def is_rate_limited(exc: Exception) -> bool:
    text = (repr(exc) + " " + str(exc)).lower()
    return "rate_limit" in text or "429" in text or "rate limit" in text


def is_moderation_block(exc: Exception) -> bool:
    text = (repr(exc) + " " + str(exc)).lower()
    return "moderation_blocked" in text or "safety system" in text


def call_openai_image(client, prompt: str, *, model: str, size: str,
                      quality: str, background: str,
                      max_retries: int, retry_wait: float,
                      download_timeout: float) -> bytes:
    last_exc: Optional[Exception] = None
    for attempt in range(1, max_retries + 1):
        try:
            kwargs = {
                "model": model,
                "prompt": prompt,
                "size": size,
                "n": 1,
            }
            if model == "gpt-image-1":
                kwargs["quality"] = quality
                if background and background != "auto":
                    kwargs["background"] = background
            resp = client.images.generate(**kwargs)
            datum = resp.data[0]
            b64 = getattr(datum, "b64_json", None)
            if b64:
                return base64.b64decode(b64)
            url = getattr(datum, "url", None)
            if url:
                import urllib.request
                with urllib.request.urlopen(url, timeout=download_timeout) as r:
                    return r.read()
            raise RuntimeError("OpenAI response had neither b64_json nor url")
        except Exception as exc:  # noqa: BLE001
            last_exc = exc
            if is_moderation_block(exc):
                raise
            if attempt == max_retries:
                raise
            wait = retry_wait
            if is_rate_limited(exc):
                wait = max(retry_wait, 15.0)
            elif is_transient_network_error(exc):
                wait = max(retry_wait, 5.0)
            wait *= 1 + (attempt - 1) * 0.5
            wait += random.uniform(0, 1.0)
            print(f"    [retry {attempt}/{max_retries}] {type(exc).__name__}: "
                  f"{str(exc)[:160]}  (sleep {wait:.1f}s)")
            time.sleep(wait)
    raise last_exc or RuntimeError("unknown failure")


# ---- Cost estimate ---------------------------------------------------------


def estimate_cost(num: int, model: str, quality: str) -> float:
    if model != "gpt-image-1":
        return 0.04 * num  # rough fallback for dall-e-3 standard
    per = {
        "low": 0.011,
        "medium": 0.04,
        "high": 0.17,
        "standard": 0.04,
        "auto": 0.04,
    }.get(quality, 0.04)
    return per * num


# ---- Main ------------------------------------------------------------------


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    default_input = REPO_ROOT / "tools" / "image-prompts.xlsx"
    if not default_input.is_file():
        default_input = REPO_ROOT / "tools" / "image-prompts.csv"
    parser.add_argument("--csv", type=Path,
                        default=default_input,
                        help="Input CSV or XLSX of prompts (default: image-prompts.xlsx "
                             "if present, otherwise image-prompts.csv).")
    parser.add_argument("--word-sets", type=Path,
                        default=REPO_ROOT / "word-lists" / "word-sets.json")
    parser.add_argument("--word-images-root", type=Path,
                        default=REPO_ROOT / "word-images")
    parser.add_argument("--library-dir", type=Path,
                        default=REPO_ROOT / "word-images" / "_library")
    parser.add_argument("--mirror-sets", action="store_true",
                        help="Also copy each image into per-set folders (legacy). Default: _library/ only.")
    parser.add_argument("--model", default="gpt-image-1")
    parser.add_argument("--size", default="1024x1024",
                        choices=["1024x1024", "1024x1536", "1536x1024", "auto"])
    parser.add_argument("--quality", default="medium",
                        choices=["low", "medium", "high", "auto", "standard"])
    parser.add_argument("--background", default="auto",
                        choices=["auto", "opaque", "transparent"])
    parser.add_argument("--prompt-mode", default="csv",
                        choices=["merge", "csv", "template"],
                        help="csv: use the prompt column verbatim (default). "
                             "merge: combine custom subject with the tuned template. "
                             "template: ignore CSV prompt and use the default template.")
    parser.add_argument("--concurrency", type=int, default=1)
    parser.add_argument("--min-seconds-between-jobs", type=float, default=12.0,
                        help="Sleep this long between jobs (default 12s — keeps you "
                             "under the 5/min gpt-image-1 limit).")
    parser.add_argument("--limit", type=int, default=0,
                        help="Stop after this many words (0 = no limit).")
    parser.add_argument("--only-word", action="append", default=[],
                        help="Only generate this word (can be repeated).")
    parser.add_argument("--skip-existing", action="store_true", default=True,
                        help="Skip words that already have a PNG in _library/.")
    parser.add_argument("--no-skip-existing", dest="skip_existing", action="store_false")
    parser.add_argument("--skip-highlighted", action="store_true", default=True,
                        help="When reading .xlsx, skip rows with any cell highlighted.")
    parser.add_argument("--no-skip-highlighted", dest="skip_highlighted",
                        action="store_false")
    parser.add_argument("--dry-run", action="store_true",
                        help="Don't call OpenAI; just print what would happen.")
    parser.add_argument("--stop-on-error", action="store_true")
    parser.add_argument("--max-retries", type=int, default=4)
    parser.add_argument("--retry-wait-seconds", type=float, default=5.0)
    parser.add_argument("--download-timeout", type=float, default=60.0)
    parser.add_argument(
        "--request-timeout",
        type=float,
        default=180.0,
        help="Abort a single OpenAI API call after this many seconds (default 180). "
             "Prevents the script from hanging silently for hours.",
    )
    return parser.parse_args()


def main() -> int:
    load_dotenv_simple(REPO_ROOT / ".env")
    args = parse_args()

    rows = load_prompt_rows(args.csv, skip_highlighted=args.skip_highlighted)
    if args.only_word:
        wanted = {w.lower() for w in args.only_word}
        rows = [r for r in rows if r["word"].lower() in wanted]

    sets = load_word_sets(args.word_sets)
    folder_map = word_to_folders(sets)

    args.library_dir.mkdir(parents=True, exist_ok=True)

    if args.skip_existing:
        before = len(rows)
        rows = [r for r in rows
                if not (args.library_dir / f"{r['word']}.png").is_file()]
        skipped = before - len(rows)
        if skipped:
            print(f"Skipping {skipped} word(s) already in _library/.")

    if args.limit > 0:
        rows = rows[:args.limit]

    if not rows:
        print("Nothing to do — no rows to generate.")
        return 0

    cost = estimate_cost(len(rows), args.model, args.quality)
    print(f"Will generate {len(rows)} image(s) with {args.model} ({args.quality}).")
    print(f"Estimated cost: ~${cost:.2f}")
    print(f"Output: {args.library_dir}" +
          (" + per-set folders" if args.mirror_sets else ""))
    print(f"Pacing: {args.min_seconds_between_jobs:.1f}s between jobs, "
          f"concurrency={args.concurrency}, retries={args.max_retries}, "
          f"request_timeout={args.request_timeout:.0f}s")
    print()

    if args.dry_run:
        for r in rows[:25]:
            final = build_final_prompt(r["word"], r["prompt"], args.prompt_mode)
            print(f"  [dry] {r['word']}: {final[:120]}")
        if len(rows) > 25:
            print(f"  ... plus {len(rows) - 25} more.")
        return 0

    try:
        from openai import OpenAI
    except ImportError:
        raise SystemExit("Install the OpenAI Python SDK: pip3 install openai")
    if not os.environ.get("OPENAI_API_KEY"):
        raise SystemExit(
            "OPENAI_API_KEY is not set. Add it to your .env file:\n"
            f"    {REPO_ROOT / '.env'}\n"
            "(One line: OPENAI_API_KEY=sk-...)"
        )
    client = OpenAI(timeout=args.request_timeout)

    ok, fail, last_job_time = 0, 0, 0.0
    total = len(rows)

    def run_one(row: Dict[str, str], index: int) -> Tuple[str, bool, str]:
        word = row["word"]
        final_prompt = build_final_prompt(word, row["prompt"], args.prompt_mode)
        print(f"  [{index}/{total}] Generating {word}...", flush=True)
        try:
            data = call_openai_image(
                client, final_prompt,
                model=args.model, size=args.size, quality=args.quality,
                background=args.background,
                max_retries=args.max_retries,
                retry_wait=args.retry_wait_seconds,
                download_timeout=args.download_timeout,
            )
            paths = destination_paths(
                word, row["folder"], folder_map,
                args.word_images_root, args.library_dir,
                mirror_sets=args.mirror_sets,
            )
            n = write_png_bytes(paths, data)
            return (word, True, f"wrote {n} file(s)")
        except Exception as exc:  # noqa: BLE001
            return (word, False, f"{type(exc).__name__}: {str(exc)[:200]}")

    if args.concurrency <= 1:
        for i, r in enumerate(rows, start=1):
            now = time.monotonic()
            gap = args.min_seconds_between_jobs - (now - last_job_time)
            if gap > 0:
                time.sleep(gap)
            last_job_time = time.monotonic()
            word, success, msg = run_one(r, i)
            tag = "ok" if success else "fail"
            print(f"  [{tag}] {word}: {msg}", flush=True)
            if success:
                ok += 1
            else:
                fail += 1
                if args.stop_on_error:
                    break
    else:
        with ThreadPoolExecutor(max_workers=args.concurrency) as pool:
            futures = []
            for i, r in enumerate(rows, start=1):
                now = time.monotonic()
                gap = args.min_seconds_between_jobs - (now - last_job_time)
                if gap > 0:
                    time.sleep(gap)
                last_job_time = time.monotonic()
                futures.append(pool.submit(run_one, r, i))
            for fut in as_completed(futures):
                word, success, msg = fut.result()
                tag = "ok" if success else "fail"
                print(f"  [{tag}] {word}: {msg}", flush=True)
                if success:
                    ok += 1
                else:
                    fail += 1
                    if args.stop_on_error:
                        for f in futures:
                            f.cancel()
                        break

    print()
    print(f"Done. ok={ok} fail={fail}")
    return 0 if fail == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
