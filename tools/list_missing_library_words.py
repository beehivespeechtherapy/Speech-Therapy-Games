#!/usr/bin/env python3
"""List every word in word-sets.json that has no matching image in _library/.

Includes words from minimal-pair sets (``pairs``) and single-word sets (``words``).

This is the practical definition of "still need a clipart" (we cannot detect
true "deletions" from disk history). Output files are written under tools/:

  missing-library-words.csv   — columns: word, example_folder
  missing-library-words.txt   — one word per line (easy to skim)

Optional: compare against image-prompts.xlsx (non-highlighted rows only) and
also write:

  missing-not-in-prompts-xlsx.csv

Usage (from project root):

    python3 tools/list_missing_library_words.py
    python3 tools/list_missing_library_words.py --no-xlsx-compare
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from pathlib import Path
from typing import Dict, Iterable, Set, Tuple


def find_repo_root(start: Path) -> Path:
    here = start.resolve()
    for candidate in [here, *here.parents]:
        if (candidate / "word-lists").is_dir() and (candidate / "word-images").is_dir():
            return candidate
    return start.resolve()


REPO_ROOT = find_repo_root(Path(__file__).parent)


def cell_is_highlighted(cell) -> bool:
    try:
        fill = cell.fill
        if not fill or not getattr(fill, "patternType", None) or fill.patternType == "none":
            return False
        color = fill.fgColor
        ctype = getattr(color, "type", None)
        if ctype == "theme":
            theme = getattr(color, "theme", None)
            tint = getattr(color, "tint", None) or 0.0
            if (theme == 0 or theme is None) and abs(tint) < 1e-6:
                return False
            return True
        if ctype == "rgb":
            rgb = getattr(color, "rgb", None)
            if isinstance(rgb, str) and rgb.upper() not in {"00000000", "FFFFFFFF"}:
                return True
            return False
        if ctype == "indexed":
            return getattr(color, "indexed", None) not in (None, 64, 65)
        return True
    except Exception:
        return False


def load_xlsx_words(path: Path) -> Set[str]:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    header_cells = next(ws.iter_rows(max_row=1), None)
    if not header_cells:
        return set()
    header = [(str(c.value).strip().lower() if c.value else "") for c in header_cells]
    idx_word = header.index("word")
    out: Set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=False):
        if not row or len(row) <= idx_word:
            continue
        if any(cell_is_highlighted(c) for c in row):
            continue
        v = row[idx_word].value
        if v:
            out.add(str(v).strip().lower())
    return out


def iter_words_in_set(word_set: dict) -> Iterable[Tuple[str, str]]:
    """Yield (word, folder) for every word in a set (minimal pairs or single-word lists)."""
    folder = (word_set.get("folder") or "").strip()
    if not folder:
        return
    set_type = (word_set.get("setType") or "pairs").strip().lower()
    if set_type == "single":
        for raw in word_set.get("words", []) or []:
            w = str(raw).strip()
            if w:
                yield w, folder
        return
    for pair in word_set.get("pairs", []) or []:
        for raw in pair:
            w = str(raw).strip()
            if w:
                yield w, folder


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--word-sets", type=Path, default=REPO_ROOT / "word-lists" / "word-sets.json")
    parser.add_argument("--library-dir", type=Path, default=REPO_ROOT / "word-images" / "_library")
    parser.add_argument("--output-dir", type=Path, default=REPO_ROOT / "tools")
    parser.add_argument("--prompt-xlsx", type=Path, default=REPO_ROOT / "tools" / "image-prompts.xlsx")
    parser.add_argument("--no-xlsx-compare", action="store_true")
    args = parser.parse_args()

    with args.word_sets.open(encoding="utf-8") as f:
        sets = json.load(f)

    lib_stems: Set[str] = set()
    if args.library_dir.is_dir():
        for p in args.library_dir.glob("*.png"):
            key = p.stem.strip().lower()
            if key:
                lib_stems.add(key)

    lower_to_display: Dict[str, str] = {}
    folders_for: Dict[str, Set[str]] = {}
    pair_sets = 0
    single_sets = 0

    for s in sets:
        set_type = (s.get("setType") or "pairs").strip().lower()
        if set_type == "single":
            single_sets += 1
        else:
            pair_sets += 1
        for w, folder in iter_words_in_set(s):
            wl = w.lower()
            if wl not in lower_to_display:
                lower_to_display[wl] = w
            folders_for.setdefault(wl, set()).add(folder)

    missing_lower = sorted(set(lower_to_display) - lib_stems, key=lambda k: lower_to_display[k].lower())

    out_dir = args.output_dir
    out_dir.mkdir(parents=True, exist_ok=True)
    csv_path = out_dir / "missing-library-words.csv"
    txt_path = out_dir / "missing-library-words.txt"

    with csv_path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["word", "example_folder"])
        for wl in missing_lower:
            ex = sorted(folders_for[wl])[0] if folders_for.get(wl) else ""
            w.writerow([lower_to_display[wl], ex])

    with txt_path.open("w", encoding="utf-8") as fh:
        fh.write(
            "# Words in word-lists/word-sets.json with NO *.png in word-images/_library/\n"
            "# (case-insensitive filename stem). Not the same as 'provably deleted files'.\n"
            f"# Total: {len(missing_lower)}\n\n"
        )
        for wl in missing_lower:
            fh.write(lower_to_display[wl] + "\n")

    print(f"Scanned {len(sets)} set(s): {pair_sets} pair, {single_sets} single-word.")
    print(f"Unique words in word-sets.json: {len(lower_to_display)}")
    print(f"Words with no _library PNG: {len(missing_lower)}")
    print(f"Wrote: {csv_path}")
    print(f"Wrote: {txt_path}")

    if not args.no_xlsx_compare and args.prompt_xlsx.is_file():
        in_xlsx = load_xlsx_words(args.prompt_xlsx)
        not_in_xlsx = [wl for wl in missing_lower if wl not in in_xlsx]
        extra_path = out_dir / "missing-not-in-prompts-xlsx.csv"
        with extra_path.open("w", newline="", encoding="utf-8") as fh:
            w = csv.writer(fh)
            w.writerow(["word", "example_folder"])
            for wl in not_in_xlsx:
                ex = sorted(folders_for[wl])[0] if folders_for.get(wl) else ""
                w.writerow([lower_to_display[wl], ex])
        already = len(missing_lower) - len(not_in_xlsx)
        print()
        print(f"Compared to non-highlighted rows in {args.prompt_xlsx.name}:")
        print(f"  Missing from _library but already on sheet: {already}")
        print(f"  Missing from _library and NOT on sheet (add these): {len(not_in_xlsx)}")
        print(f"Wrote: {extra_path}")
        print("  (word list only — run: python3 tools/generate_missing_prompt_sheet.py for prompts.)")

    return 0


if __name__ == "__main__":
    sys.exit(main())
